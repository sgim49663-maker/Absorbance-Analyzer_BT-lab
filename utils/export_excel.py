import io
import pandas as pd
from utils import natural_sort_key
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.series import DataPoint
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter
from openpyxl.chart.layout import Layout, ManualLayout

def generate_excel(result: dict) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        stats_df = result["stats_df"].copy()
        
        rename_cols = {}
        if "Mean_Value" in stats_df.columns:
            if result.get("assay_type") in ["ELISA", "Fluorescence"]:
                rename_cols["Mean_Value"] = "Mean (Conc)"
                rename_cols["SD_Value"] = "SD (Conc)"
            elif result.get("assay_type") == "DPPH":
                rename_cols["Mean_Value"] = "Activity (%)"
                rename_cols["SD_Value"] = "Activity SD"
            else:
                rename_cols["Mean_Value"] = "Viability (%)"
                rename_cols["SD_Value"] = "Viability SD"
        
        stats_df.rename(columns=rename_cols, inplace=True)
        stats_df.to_excel(writer, sheet_name="Statistics", index=False)

        proc_df = result["processed_df"].copy()
        proc_df.to_excel(writer, sheet_name="Processed Data", index=False, startrow=12)

        wb = writer.book
        ws_stat = wb["Statistics"]

        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for cell in ws_stat[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row in ws_stat.iter_rows(min_row=2, max_row=ws_stat.max_row, max_col=ws_stat.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                col_name = ws_stat.cell(row=1, column=cell.column).value
                if isinstance(cell.value, float):
                    if col_name and str(col_name).startswith("p vs"):
                        cell.number_format = "0.00E+00"
                    else:
                        cell.number_format = "0.00"

        for i, cell in enumerate(ws_stat[1], 1):
            if cell.value and str(cell.value).startswith("Sig vs"):
                for row in ws_stat.iter_rows(min_row=2, max_row=ws_stat.max_row,
                                              min_col=i, max_col=i):
                    for c in row:
                        if c.value and "*" in str(c.value):
                            c.font = Font(color="FF0000", bold=True, size=12)

        for col_cells in ws_stat.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells) + 3
            ws_stat.column_dimensions[col_cells[0].column_letter].width = max_len

        ws_info = wb.create_sheet("Analysis Info")
        info_data = [
            ["Parameter", "Value"],
            ["Assay Type", result.get("assay_type", "N/A")],
            ["Blank Mean", result.get("bl_mean", 0)],
            ["1st Reference", result.get("ref_1st", "N/A")],
            ["2nd Reference", result.get("ref_2nd", "N/A")],
        ]
        if result.get("elisa_curve"):
            curve = result["elisa_curve"]
            if not curve.get("error"):
                info_data.append(["Curve Fit", curve.get("curve_fit", "Linear")])
                info_data.append(["Equation", curve["equation"]])
                info_data.append(["R-squared", curve["r_squared"]])
                if curve.get("curve_fit") == "4PL":
                    info_data.append(["A (Min asymptote)", curve.get("A", "")])
                    info_data.append(["B (Hill slope)", curve.get("B", "")])
                    info_data.append(["C (Inflection/EC50)", curve.get("C", "")])
                    info_data.append(["D (Max asymptote)", curve.get("D", "")])
                
        for r_idx, row_data in enumerate(info_data, 1):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_info.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        ws_proc = wb["Processed Data"]
        rows = list("ABCDEFGH")
        
        ws_proc.cell(row=1, column=1, value="Plate Map Visualization").font = Font(bold=True, size=12)
        
        well_dict = {}
        for _, r in proc_df.iterrows():
            well_dict[r["Well"]] = {"label": r["Label"], "type": r["Type"], "excluded": r.get("Excluded", False)}
            
        color_map = {
            "BL": "E0E0E0",
            "ST": "FFE0B2",
            "NC": "FFFFFF",
            "PC": "BDBDBD",
            "SM": "C5E1A5",
        }
        
        for c in range(1, 13):
            cell = ws_proc.cell(row=3, column=c+1, value=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        for r_idx, row_chr in enumerate(rows):
            cell_row = ws_proc.cell(row=r_idx+4, column=1, value=row_chr)
            cell_row.font = Font(bold=True)
            cell_row.alignment = Alignment(horizontal="center")
            
            for c in range(1, 13):
                wid = f"{row_chr}{c}"
                target_cell = ws_proc.cell(row=r_idx+4, column=c+1)
                
                target_cell.value = ""
                c_fill = "D9D9D9"
                font_clr = "808080"
                
                if wid in well_dict:
                    info = well_dict[wid]
                    c_fill = color_map.get(info["type"], "FFFFFF")
                    if info["excluded"]:
                        c_fill = "D3D3D3"
                        font_clr = "808080"
                        target_cell.value = f"[{info['label']}]"
                    else:
                        target_cell.value = info["label"]
                        font_clr = "000000"
                
                target_cell.fill = PatternFill(start_color=c_fill, end_color=c_fill, fill_type="solid")
                target_cell.font = Font(color=font_clr)
                target_cell.border = thin_border
                target_cell.alignment = Alignment(horizontal="center")

        col_count = len(proc_df.columns)
        for c_idx in range(1, col_count + 1):
            c = ws_proc.cell(row=13, column=c_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
            
        _add_editable_charts(wb, result)

    output.seek(0)
    return output.read()


def _add_editable_charts(wb, result):
    ws = wb["Statistics"]
    stats_df = result["stats_df"].copy()
    ref_1st = result.get("ref_1st")
    custom_names = result.get("custom_names", {})
    assay_type = result.get("assay_type", "Cell Viability")

    order_labels = []
    if ref_1st and ref_1st in stats_df["Label"].values:
        order_labels.append(ref_1st)
    remaining = sorted(
        [lab for lab in stats_df["Label"].unique() if lab not in order_labels],
        key=natural_sort_key,
    )
    order_labels.extend(remaining)

    stats_df["_order"] = stats_df["Label"].apply(lambda x: order_labels.index(x) if x in order_labels else 999)
    stats_df = stats_df.sort_values("_order").reset_index(drop=True)

    data_start_col = ws.max_column + 2
    row_offset = 2
    
    ws.cell(row=row_offset-1, column=data_start_col, value="Sample")
    ws.cell(row=row_offset-1, column=data_start_col+1, value="Mean")
    ws.cell(row=row_offset-1, column=data_start_col+2, value="SD")

    sig_cols = [c for c in stats_df.columns if str(c).startswith("Sig vs") or str(c).startswith("p vs")]
    
    # Sort nicely if needed or just use as is, let's group p vs and Sig vs for the same ref together.
    # Actually, stats_df.columns already has them in order: p vs Ref1, Sig vs Ref1, p vs Ref2, Sig vs Ref2
    extra_cols = [c for c in stats_df.columns if str(c).startswith("p vs") or str(c).startswith("Sig vs")]
    
    for j, c_name in enumerate(extra_cols):
        ws.cell(row=row_offset-1, column=data_start_col+3+j, value=c_name)

    n_samples = len(stats_df)
    mean_col = "Mean_Value" if "Mean_Value" in stats_df.columns else "Mean (%)"
    sd_col = "SD_Value" if "SD_Value" in stats_df.columns else "SD (%)"

    for i, row in stats_df.iterrows():
        r = i + row_offset
        label = row["Label"]
        display = custom_names.get(label, label)
        ws.cell(row=r, column=data_start_col, value=display)
        ws.cell(row=r, column=data_start_col+1, value=row.get(mean_col, 0))
        ws.cell(row=r, column=data_start_col+2, value=row.get(sd_col, 0))
        for j, c_name in enumerate(extra_cols):
            ws.cell(row=r, column=data_start_col+3+j, value=row.get(c_name, "-"))
        
    ws.column_dimensions[get_column_letter(data_start_col)].width = 15
    ws.column_dimensions[get_column_letter(data_start_col+1)].width = 10
    ws.column_dimensions[get_column_letter(data_start_col+2)].width = 10
    for j in range(len(extra_cols)):
        ws.column_dimensions[get_column_letter(data_start_col+3+j)].width = 12


    # ── Bar Chart ──
    chart = BarChart()
    chart.type = "col"
    chart.title = None
    chart.y_axis.majorGridlines = None
    
    if assay_type == "Cell Viability": chart.y_axis.title = "Cell Viability (%)"
    elif assay_type == "DPPH": chart.y_axis.title = "DPPH radical scavenging activity (%)"
    elif assay_type in ["ELISA", "Fluorescence"]: chart.y_axis.title = "Concentration"
    else: chart.y_axis.title = "Relative Absorbance (%)"
        
    chart.x_axis.title = None
    chart.legend = None

    # 실선 x, y 축 테두리 및 엑셀 축 영역 강제 보임 활성화
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="000000", w=6350))
    chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="000000", w=6350))
    chart.y_axis.scaling.min = 0
    chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.05, w=0.92, h=0.8))

    data_ref = Reference(ws, min_col=data_start_col+1, min_row=row_offset-1, max_row=n_samples + row_offset - 1)
    cats_ref = Reference(ws, min_col=data_start_col, min_row=row_offset, max_row=n_samples + row_offset - 1)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    series = chart.series[0]

    # 양방향 오차 막대
    series.errBars = ErrorBars(errBarType="both", errValType="cust", noEndCap=False)
    
    from openpyxl.chart.series import NumRef, NumDataSource
    err_let = get_column_letter(data_start_col+2)
    num_ref = NumRef(f"'Statistics'!${err_let}${row_offset}:${err_let}${n_samples + row_offset - 1}")
    series.errBars.plus = NumDataSource(numRef=num_ref)

    labels_list = stats_df["Label"].tolist()
    for idx, lab in enumerate(labels_list):
        pt = DataPoint(idx=idx)
        _set_bar_fill(pt, "FFFFFF", "000000")
        series.data_points.append(pt)

    # (Baseline Line Chart Removed)

    chart.width = 16
    chart.height = 10

    insert_row = ws.max_row + 2
    ws.add_chart(chart, f"A{insert_row}")

    # ── Standard Curve (Editable Scatter Chart) ──
    if assay_type in ["ELISA", "Fluorescence"] and result.get("elisa_curve") and not result["elisa_curve"].get("error"):
        st_concs = result.get("st_concs", {})
        proc_df = result["processed_df"]
        curve_info = result["elisa_curve"]
        
        # 평균낸 ST points로 추출
        st_groups = {}
        for _, r in proc_df.iterrows():
            if r["Type"] == "ST" and r["Label"] in st_concs and pd.notna(r["BL_corrected"]):
                st_groups.setdefault(r["Label"], []).append(r["BL_corrected"])
                
        st_data = []
        for lab, vals in st_groups.items():
            st_data.append((st_concs[lab], float(sum(vals)/len(vals))))
                
        if st_data:
            # Write ST data to worksheet for the chart
            st_col = data_start_col + 5
            ws.cell(row=row_offset-1, column=st_col, value="ST Conc")
            ws.cell(row=row_offset-1, column=st_col+1, value="Absorbance")
            
            for i, (x_val, y_val) in enumerate(st_data):
                ws.cell(row=row_offset+i, column=st_col, value=x_val)
                ws.cell(row=row_offset+i, column=st_col+1, value=y_val)
                
            ws.column_dimensions[get_column_letter(st_col)].width = 12
            ws.column_dimensions[get_column_letter(st_col+1)].width = 12
            
            scatter = ScatterChart()
            scatter.title = "Standard Curve"
            scatter.x_axis.title = 'Concentration'
            scatter.y_axis.title = 'Absorbance (BL Corrected)'
            
            # 실선 x, y 축 강제 표출
            scatter.x_axis.delete = False
            scatter.y_axis.delete = False
            scatter.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="000000", w=6350))
            scatter.y_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="000000", w=6350))
            scatter.y_axis.scaling.min = 0
            scatter.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.05, w=0.92, h=0.8))
            
            xvalues = Reference(ws, min_col=st_col, min_row=row_offset, max_row=row_offset + len(st_data) - 1)
            yvalues = Reference(ws, min_col=st_col+1, min_row=row_offset-1, max_row=row_offset + len(st_data) - 1)
            
            s2 = Series(yvalues, xvalues, title_from_data=True)
            s2.marker.symbol = "circle"
            s2.marker.graphicalProperties.solidFill = "333333"
            s2.marker.graphicalProperties.ln = LineProperties(solidFill="333333")
            s2.graphicalProperties.ln = LineProperties(noFill=True)  # Hide connecting line
            
            scatter.series.append(s2)
            
            if curve_info.get("curve_fit") == "4PL":
                # ── 4PL: manually compute curve data and add as line series ──
                from utils.analysis import logistic4
                import numpy as np
                
                A = curve_info["A"]
                B = curve_info["B"]
                C = curve_info["C"]
                D = curve_info["D"]
                
                x_data = [d[0] for d in st_data]
                positive_x = [v for v in x_data if v > 0]
                if positive_x:
                    x_min_val = min(positive_x) * 0.5
                    x_max_val = max(positive_x) * 2.0
                else:
                    x_min_val, x_max_val = 0.1, 100
                
                # Generate curve points
                curve_x = np.logspace(np.log10(x_min_val), np.log10(x_max_val), 50)
                curve_y = logistic4(curve_x, A, B, C, D)
                
                # Write curve data to sheet
                curve_col = st_col + 3
                ws.cell(row=row_offset-1, column=curve_col, value="Curve X")
                ws.cell(row=row_offset-1, column=curve_col+1, value="4PL Curve")
                ws.column_dimensions[get_column_letter(curve_col)].width = 12
                ws.column_dimensions[get_column_letter(curve_col+1)].width = 12
                
                for i, (cx, cy) in enumerate(zip(curve_x, curve_y)):
                    ws.cell(row=row_offset+i, column=curve_col, value=float(cx))
                    ws.cell(row=row_offset+i, column=curve_col+1, value=float(cy))
                
                # Add curve line series
                curve_x_ref = Reference(ws, min_col=curve_col, min_row=row_offset, max_row=row_offset + len(curve_x) - 1)
                curve_y_ref = Reference(ws, min_col=curve_col+1, min_row=row_offset-1, max_row=row_offset + len(curve_x) - 1)
                
                s_curve = Series(curve_y_ref, curve_x_ref, title_from_data=True)
                s_curve.graphicalProperties.ln = LineProperties(solidFill="FF0000", prstDash="dash", w=19050)
                s_curve.marker.symbol = "none"
                s_curve.smooth = True
                
                scatter.series.append(s_curve)
                
                # Log scale X axis for 4PL
                scatter.x_axis.scaling.logBase = 10
                
                # Add equation annotation in title
                eq_str = curve_info["equation"]
                r2_str = f"R² = {curve_info['r_squared']:.4f}"
                scatter.title = f"Standard Curve\n{eq_str}\n{r2_str}"
                
            else:
                # ── Linear: use trendline ──
                from openpyxl.chart.trendline import Trendline
                s2.trendline = Trendline(trendlineType="linear", dispEq=True, dispRSqr=True)
                s2.trendline.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="FF0000", prstDash="dash"))
            
            scatter.legend = None
            scatter.width = 14
            scatter.height = 10
            
            ws.add_chart(scatter, f"H{insert_row}")

def _set_bar_fill(data_point, fill_color, border_color):
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    gp = GraphicalProperties()
    gp.solidFill = fill_color
    gp.ln = LineProperties(solidFill=border_color, w=9525)
    data_point.graphicalProperties = gp


